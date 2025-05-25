



# views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db import IntegrityError
from .forms import UserLoginForm
from django.contrib.auth import authenticate, login
from .models import *
import os
import re
import cv2
import pytesseract
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.urls import reverse
from django.conf import settings
from django.http import FileResponse
from django.shortcuts import render, redirect
from django.core.files.storage import FileSystemStorage


def loginpage(request):
    if request.method == 'POST':
        form = UserLoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('index')  # Redirect to a home page or any other page
    else:
        form = UserLoginForm()
    return render(request, 'core/userlog.html', {'form': form})

def index(request):
    return render(request, 'core/index.html')

UPLOAD_FOLDER = os.path.join(settings.MEDIA_ROOT, 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def preprocess_image(image_path):
    image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    _, binary_image = cv2.threshold(image, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    return binary_image

def perform_ocr(image):
    custom_config = r'--oem 3 --psm 6'
    return pytesseract.image_to_string(image, config=custom_config)

def interpret_and_structure_text(text):
    lines = text.split('\n')
    headings = {}
    pattern = re.compile(r'^(.*?):\s*(.*)$')

    for line in lines:
        match = pattern.match(line.strip())
        if match:
            heading, value = match.groups()
            if heading not in headings:
                headings[heading] = []
            headings[heading].append(value)

    max_length = max(len(values) for values in headings.values()) if headings else 0
    structured_data = [
        [values[i] if i < len(values) else '' for values in headings.values()]
        for i in range(max_length)
    ]
    return list(headings.keys()), structured_data

def create_or_append_excel(headings, data, output_excel):
    df_new = pd.DataFrame(data, columns=headings)
    if os.path.exists(output_excel):
        book = load_workbook(output_excel)
        if 'Sheet1' in book.sheetnames:
            ws = book['Sheet1']
            for row in dataframe_to_rows(df_new, index=False, header=False):
                ws.append(row)
            book.save(output_excel)
        else:
            with pd.ExcelWriter(output_excel, engine='openpyxl', mode='a') as writer:
                df_new.to_excel(writer, index=False, sheet_name='Sheet1')
    else:
        df_new.to_excel(output_excel, index=False, sheet_name='Sheet1')


def upload_file(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        fs = FileSystemStorage(location=UPLOAD_FOLDER)
        file_path = fs.save(uploaded_file.name, uploaded_file)
        full_path = os.path.join(UPLOAD_FOLDER, file_path)
        process_image(full_path)
        return redirect(reverse('index') + '?uploaded=true')
    return redirect('index')

def process_image(image_path):
    preprocessed = preprocess_image(image_path)
    text = perform_ocr(preprocessed)
    headings, data = interpret_and_structure_text(text)
    output_excel = os.path.join(settings.BASE_DIR, 'output.xlsx')
    create_or_append_excel(headings, data, output_excel)

def download_file(request):
    output_excel = os.path.join(settings.BASE_DIR, 'output.xlsx')
    if os.path.exists(output_excel):
        return FileResponse(open(output_excel, 'rb'), as_attachment=True)
    return redirect('core/index')
